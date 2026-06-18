from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl


CODE_HEADER = "證券代號"
NAME_HEADER = "證券名稱"


def main() -> None:
    parser = argparse.ArgumentParser(description="Build stock catalog JSON from an xlsx stock list.")
    parser.add_argument("input", type=Path, help="Input .xlsx path.")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/stock_catalog.json"),
        help="Output JSON path. Default: data/stock_catalog.json",
    )
    args = parser.parse_args()

    items = extract_items(args.input)
    payload = {
        "source": str(args.input),
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "count": len(items),
        "items": items,
    }

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"Wrote {len(items)} stock catalog entries to {args.output}")


def extract_items(path: Path) -> list[dict[str, str]]:
    if not path.is_file():
        raise FileNotFoundError(path)

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    seen: set[str] = set()
    items: list[dict[str, str]] = []

    for sheet in workbook.worksheets:
        header = _find_header(sheet)
        if header is None:
            continue
        header_row, code_col, name_col = header
        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            stock_id = _cell_text(_safe_cell(row, code_col))
            name = _cell_text(_safe_cell(row, name_col))
            if not stock_id or not name:
                continue
            if not _looks_like_stock_id(stock_id):
                continue
            if stock_id in seen:
                continue
            seen.add(stock_id)
            items.append(
                {
                    "stock_id": stock_id,
                    "market": "TWSE",
                    "name": name,
                    "short_name": name,
                }
            )

    if not items:
        raise ValueError(f"No stock rows found in {path}")
    return items


def _find_header(sheet: Any) -> tuple[int, int, int] | None:
    max_scan_rows = min(sheet.max_row or 0, 20)
    for row_index, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True),
        start=1,
    ):
        values = [_cell_text(value) for value in row]
        try:
            code_col = values.index(CODE_HEADER)
            name_col = values.index(NAME_HEADER)
        except ValueError:
            continue
        return row_index, code_col, name_col
    return None


def _safe_cell(row: tuple[Any, ...], index: int) -> Any:
    if index < 0 or index >= len(row):
        return None
    return row[index]


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def _looks_like_stock_id(value: str) -> bool:
    if not value:
        return False
    return value.isalnum() and 4 <= len(value) <= 8


if __name__ == "__main__":
    main()
