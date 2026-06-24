from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


DISCLAIMER = "本表為已同步資料之整理，非投資建議，不預測股價。"
SOURCE_NOTE = "資料來源：本機已同步資料"
BRAND_FILL = "173B57"
HEADER_FILL = "EAF1F5"
NOTE_FILL = "F7FAFC"
BORDER_COLOR = "D9E2EA"


def build_portfolio_workbook_bytes(
    payload: dict[str, Any],
    *,
    generated_at: datetime | None = None,
) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "持倉總覽"
    _write_portfolio_summary(summary, payload, generated_at=generated_at)

    transactions = workbook.create_sheet("交易紀錄")
    _write_transactions(transactions, payload, generated_at=generated_at)

    dividends = workbook.create_sheet("配息紀錄")
    _write_portfolio_dividends(dividends, payload, generated_at=generated_at)

    notes = workbook.create_sheet("說明")
    _write_notes(
        notes,
        [
            ["匯出口徑", "本表匯出目前 App 已計算出的持倉、市值、帳面損益、含息總報酬、XIRR 與交易紀錄。"],
            ["尚未納入", "股票股利、除權息成本調整、完整稅務與多帳戶尚未納入。"],
            ["資料使用", "匯出資料只留在本機產生，不會上傳。"],
        ],
        generated_at=generated_at,
    )

    return _save_workbook(workbook)


def build_stock_workbook_bytes(
    payload: dict[str, Any],
    *,
    generated_at: datetime | None = None,
) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "個股摘要"
    _write_stock_summary(summary, payload, generated_at=generated_at)

    structure = workbook.create_sheet("結構指紋")
    _write_structure_fingerprint(structure, payload, generated_at=generated_at)

    prices = workbook.create_sheet("日線資料")
    _write_prices(prices, payload, generated_at=generated_at)

    annotations = workbook.create_sheet("圖表標註")
    _write_chart_annotations(annotations, payload, generated_at=generated_at)

    dividends = workbook.create_sheet("股利資料")
    _write_dividends(dividends, payload, generated_at=generated_at)

    valuation = workbook.create_sheet("估值情境")
    _write_valuation(valuation, payload, generated_at=generated_at)

    health = workbook.create_sheet("白話健檢")
    _write_health_report(health, payload, generated_at=generated_at)

    revenue = workbook.create_sheet("月營收")
    _write_monthly_revenue(revenue, payload, generated_at=generated_at)

    financials = workbook.create_sheet("財報摘要")
    _write_financials(financials, payload, generated_at=generated_at)

    return _save_workbook(workbook)


def build_screener_workbook_bytes(
    payload: dict[str, Any],
    *,
    generated_at: datetime | None = None,
) -> bytes:
    """把雷達中心（全市場批次）的完整資訊倒成 Excel，供進一步篩選/複核。"""
    generated_at = generated_at or _parse_generated_at(payload.get("generated_at"))
    workbook = Workbook()
    overview = workbook.active
    overview.title = "雷達總表"
    _write_screener_table(overview, payload, generated_at=generated_at)

    legacy = workbook.create_sheet("傳統估價表")
    _write_legacy_valuation_table(legacy, payload, generated_at=generated_at)

    notes = workbook.create_sheet("說明")
    targets = payload.get("targets") or {}
    _write_notes(
        notes,
        [
            ["這份是什麼", "雷達中心『全市場批次掃描』的完整資料表，方便自己排序、過濾、複核。"],
            ["情境價怎麼算", "以『近 5 年平均現金股利 ÷ 目標殖利率』反推的參考情境，不代表股票真正值多少，也不是買賣點。"],
            ["三個目標殖利率", f"高殖利率情境 {targets.get('cheap', 6.25)}%、中殖利率情境 {targets.get('fair', 5.0)}%、低殖利率情境 {targets.get('expensive', 3.125)}%。"],
            ["怎麼濾掉地雷", "看『股利法適用狀態』與『疑似殖利率陷阱』兩欄：不適合或低信心、或標為陷阱者，情境參考性低。"],
            ["紅線", "本表只整理已同步的公開資料，不報明牌、不預測股價、不構成投資建議。"],
        ],
        generated_at=generated_at,
    )
    return _save_workbook(workbook)


def _write_screener_table(
    sheet: Worksheet,
    payload: dict[str, Any],
    *,
    generated_at: datetime | None,
) -> None:
    _setup_sheet(sheet, "雷達總表 · 全市場批次掃描", generated_at=generated_at, last_column=23)
    summary = payload.get("summary") or {}
    overview_rows = [
        ["掃描檔數", summary.get("rows")],
        ["可估價檔數", summary.get("available_rows")],
        ["股利法適用（高信心）", summary.get("below_cheap_high_conf_rows")],
        ["高殖利率（已排陷阱）", summary.get("yield_normal_rows")],
        ["疑似殖利率陷阱", summary.get("yield_trap_rows")],
    ]
    _write_table(sheet, 4, ["統計", "檔數"], overview_rows)

    headers = [
        "代號", "名稱", "市場", "資料日", "最近收盤", "昨收", "漲跌", "漲跌%", "成交量",
        "近5年平均現金股利", "估計現金殖利率%", "股利資料年數",
        "情境價(殖利率6.25%)", "情境價(殖利率5%)", "情境價(殖利率3.125%)",
        "最近收盤相對高殖情境%", "股利法適用狀態", "公司類型", "不適用/降權原因",
        "疑似殖利率陷阱", "陷阱原因", "樣本狀態", "資料信心",
    ]
    rows: list[list[Any]] = []
    for item in payload.get("items") or []:
        rows.append(
            [
                item.get("stock_id"),
                item.get("short_name") or item.get("name"),
                item.get("market"),
                item.get("price_date"),
                item.get("latest_close", item.get("current_price")),
                item.get("previous_close"),
                item.get("day_change"),
                _as_percent_ratio(item.get("day_change_percent")),
                item.get("volume"),
                item.get("average_cash_dividend"),
                _as_percent_ratio(item.get("current_yield_percent")),
                item.get("data_years"),
                item.get("cheap_price"),
                item.get("fair_price"),
                item.get("expensive_price"),
                _as_percent_ratio(item.get("difference_percent")),
                _suitability_state_label(item.get("suitability_state")),
                item.get("company_type_label"),
                _reason_codes_label(item.get("suitability_reasons")),
                "是" if item.get("yield_trap") else "否",
                _yield_trap_label(item.get("yield_trap_reason")),
                item.get("status"),
                item.get("confidence_label"),
            ]
        )
    if not rows:
        rows = [["尚未更新雷達中心，請先在 App 按『更新雷達』"] + [None] * (len(headers) - 1)]
    _write_table(sheet, 11, headers, rows)
    last = sheet.max_row
    _format_numeric_columns(sheet, 12, last, {5, 6, 7, 10, 13, 14, 15}, "#,##0.00")
    _format_numeric_columns(sheet, 12, last, {8, 11, 16}, "0.00%")
    _format_numeric_columns(sheet, 12, last, {9, 12}, "#,##0")
    _color_screener_flags(sheet, 12, last, state_col=17, trap_col=20)
    _finish_sheet(sheet)


# 中性語意色（標資料品質，非紅綠買賣）
_STATE_FILLS = {
    "適用": "E3EEF6",
    "低信心": "FBF1D9",
    "不適合": "FBE3DD",
}
_TRAP_FILL = "FBE3DD"


def _color_screener_flags(sheet: Worksheet, start_row: int, end_row: int, *, state_col: int, trap_col: int) -> None:
    for row in range(start_row, end_row + 1):
        state_cell = sheet.cell(row, state_col)
        color = _STATE_FILLS.get(str(state_cell.value))
        if color:
            state_cell.fill = PatternFill("solid", fgColor=color)
        if str(sheet.cell(row, trap_col).value) == "是":
            sheet.cell(row, trap_col).fill = PatternFill("solid", fgColor=_TRAP_FILL)


def _recent_dividend_years(items: list[dict[str, Any]], count: int = 5) -> list[str]:
    years: set[str] = set()
    for item in items:
        by_year = item.get("annual_cash_by_year") or {}
        years.update(str(y) for y in by_year)
    return sorted(years, reverse=True)[:count]


def _legacy_flag(item: dict[str, Any]) -> str:
    if item.get("yield_trap"):
        return "⚠ 疑似殖利率陷阱"
    return _suitability_state_label(item.get("suitability_state"))


def _write_legacy_valuation_table(
    sheet: Worksheet,
    payload: dict[str, Any],
    *,
    generated_at: datetime | None,
) -> None:
    items = payload.get("items") or []
    years = _recent_dividend_years(items, count=5)
    last_column = 3 + len(years) + 6  # 代號/名稱/股價 + 各年 + 5年平均/便宜/合理/昂貴/差異/提醒
    _setup_sheet(sheet, "傳統估價表（情境參考價）", generated_at=generated_at, last_column=last_column)

    last_letter = get_column_letter(max(last_column, 2))
    sheet.merge_cells(f"A3:{last_letter}3")
    sheet["A3"] = (
        "便宜價(買價)／合理價／昂貴價(賣價) = 以近 5 年平均現金股利 ÷ 殖利率 6.25%／5%／3.125% 反推的"
        "『情境參考價』，僅供熟悉此算法者參考，非買賣建議、不預測股價；右側「提醒」標示資料品質。"
    )
    sheet["A3"].font = Font(color="8A6D1B", size=10)
    sheet["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[3].height = 30

    headers = (
        ["證券代號", "證券名稱", "最近收盤"]
        + [f"{y} 現金股利" for y in years]
        + ["5年平均", "便宜價(買價)", "合理價", "昂貴價(賣價)", "差異", "提醒"]
    )
    rows: list[list[Any]] = []
    for item in items:
        by_year = item.get("annual_cash_by_year") or {}
        rows.append(
            [
                item.get("stock_id"),
                item.get("short_name") or item.get("name"),
                item.get("latest_close", item.get("current_price")),
            ]
            + [by_year.get(str(y)) for y in years]
            + [
                item.get("average_cash_dividend"),
                item.get("cheap_price"),
                item.get("fair_price"),
                item.get("expensive_price"),
                item.get("difference"),
                _legacy_flag(item),
            ]
        )
    if not rows:
        rows = [["尚未更新雷達中心，請先在 App 按『更新雷達』"] + [None] * (len(headers) - 1)]
    _write_table(sheet, 5, headers, rows)
    last_row = sheet.max_row

    price_cols = {3} | set(range(4, 4 + len(years))) | set(range(4 + len(years), 4 + len(years) + 5))
    _format_numeric_columns(sheet, 6, last_row, price_cols, "#,##0.00")

    flag_col = len(headers)
    flag_fills = {"適用": "E3EEF6", "低信心": "FBF1D9", "不適合": "FBE3DD", "⚠ 疑似殖利率陷阱": "FBE3DD"}
    for row in range(6, last_row + 1):
        cell = sheet.cell(row, flag_col)
        color = flag_fills.get(str(cell.value))
        if color:
            cell.fill = PatternFill("solid", fgColor=color)
    _finish_sheet(sheet)


def _suitability_state_label(state: Any) -> str:
    if state == "applicable":
        return "適用"
    if state == "low_confidence":
        return "低信心"
    if state == "not_applicable":
        return "不適合"
    return str(state or "待判斷")


def _yield_trap_label(reason: Any) -> str:
    if reason == "one_off_dividend":
        return "含一次性高股利"
    if reason in (None, ""):
        return ""
    return str(reason)


_REASON_CODE_LABELS = {
    "yield_too_low": "殖利率過低",
    "loss_history": "近年曾虧損",
    "turnaround_loss": "虧損轉機",
    "insufficient_data": "股利資料不足",
    "newly_listed": "上市未滿3年",
    "unstable_dividend": "配息不穩定",
    "cyclical": "景氣循環股",
    "growth_stock": "成長股",
    "one_off_dividend": "一次性高股利",
    "high_payout": "配息率過高",
    "etf": "ETF",
    "capital_change": "資本變動",
}


def _reason_codes_label(reasons: Any) -> str:
    if not isinstance(reasons, list):
        return ""
    return "、".join(_REASON_CODE_LABELS.get(str(code), str(code)) for code in reasons)


def _parse_generated_at(value: Any) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value))
    except (TypeError, ValueError):
        return None


def _write_portfolio_summary(
    sheet: Worksheet,
    payload: dict[str, Any],
    *,
    generated_at: datetime | None,
) -> None:
    _setup_sheet(sheet, "持倉總覽", generated_at=generated_at, last_column=11)
    summary = payload.get("summary") or {}
    performance = payload.get("performance") or {}
    benchmark = performance.get("benchmark") or {}
    positions = list(payload.get("positions") or [])

    metrics = [
        ["持有檔數", summary.get("positions_count")],
        ["投入成本", summary.get("total_cost_basis")],
        ["估算市值", summary.get("total_market_value")],
        ["帳面損益", summary.get("total_unrealized_pnl")],
        ["帳面報酬率", _as_percent_ratio(summary.get("total_unrealized_return_percent"))],
        ["已實現損益", summary.get("realized_pnl")],
        ["累計現金股利", performance.get("total_cash_dividends")],
        ["含息總報酬", performance.get("total_return_amount")],
        ["含息總報酬率", _as_percent_ratio(performance.get("total_return_percent"))],
        ["年化 XIRR", _as_percent_ratio(performance.get("xirr_percent"))],
        ["0050 對比報酬率", _as_percent_ratio(benchmark.get("total_return_percent"))],
        ["成本口徑", summary.get("cost_method")],
        ["價格口徑", summary.get("price_basis")],
    ]
    _write_table(sheet, 4, ["項目", "數值"], metrics)
    for cell_ref in ["B6", "B7", "B8", "B10", "B11", "B12"]:
        sheet[cell_ref].number_format = "NT$ #,##0.00"
    for cell_ref in ["B9", "B13", "B14", "B15"]:
        sheet[cell_ref].number_format = "0.00%"

    headers = [
        "代號",
        "名稱",
        "股數",
        "平均成本",
        "成本基礎",
        "最新收盤",
        "資料日",
        "市值",
        "帳面損益",
        "報酬率",
        "佔比",
    ]
    total_market_value = _to_float(summary.get("total_market_value"))
    rows: list[list[Any]] = []
    for item in positions:
        profile = item.get("profile") or {}
        market_value = _to_float(item.get("market_value"))
        allocation = (
            market_value / total_market_value
            if total_market_value and market_value is not None
            else None
        )
        rows.append(
            [
                item.get("stock_id"),
                profile.get("short_name") or profile.get("name"),
                item.get("shares"),
                item.get("average_cost"),
                item.get("cost_basis"),
                item.get("latest_close"),
                item.get("latest_close_date"),
                item.get("market_value"),
                item.get("unrealized_pnl"),
                _as_percent_ratio(item.get("unrealized_return_percent")),
                allocation,
            ]
        )
    _write_table(sheet, 15, headers, rows or [["尚無持倉", None, None, None, None, None, None, None, None, None, None]])
    _format_numeric_columns(sheet, 16, sheet.max_row, {3, 4, 5, 6, 8, 9}, "NT$ #,##0.00")
    _format_numeric_columns(sheet, 16, sheet.max_row, {10, 11}, "0.00%")
    _finish_sheet(sheet)


def _write_transactions(
    sheet: Worksheet,
    payload: dict[str, Any],
    *,
    generated_at: datetime | None,
) -> None:
    _setup_sheet(sheet, "交易紀錄", generated_at=generated_at, last_column=9)
    headers = ["日期", "代號", "名稱", "買賣", "股數", "成交價", "手續費", "證交稅", "備註"]
    rows: list[list[Any]] = []
    for item in payload.get("transactions") or []:
        profile = item.get("profile") or {}
        rows.append(
            [
                item.get("trade_date"),
                item.get("stock_id"),
                profile.get("short_name") or profile.get("name"),
                "賣出" if item.get("side") == "sell" else "買進",
                item.get("shares"),
                item.get("price"),
                item.get("fee"),
                item.get("tax"),
                item.get("note"),
            ]
        )
    _write_table(sheet, 4, headers, rows or [["尚無交易紀錄", None, None, None, None, None, None, None, None]])
    _format_numeric_columns(sheet, 5, sheet.max_row, {5}, "#,##0")
    _format_numeric_columns(sheet, 5, sheet.max_row, {6, 7, 8}, "NT$ #,##0.00")
    _finish_sheet(sheet)


def _write_portfolio_dividends(
    sheet: Worksheet,
    payload: dict[str, Any],
    *,
    generated_at: datetime | None,
) -> None:
    _setup_sheet(sheet, "配息紀錄", generated_at=generated_at, last_column=7)
    headers = ["除息日", "代號", "當時持股", "每股現金股利", "該次領息", "來源", "備註"]
    performance = payload.get("performance") or {}
    rows = [
        [
            item.get("ex_date"),
            item.get("stock_id"),
            item.get("shares"),
            item.get("cash_dividend_per_share"),
            item.get("cash_amount"),
            item.get("source"),
            "已納入含息總報酬",
        ]
        for item in performance.get("cash_dividend_events") or []
    ]
    if not rows:
        rows = [["尚無可辨識的除息現金股利", None, None, None, None, None, None]]
    _write_table(sheet, 4, headers, rows)
    _format_numeric_columns(sheet, 5, sheet.max_row, {3}, "#,##0")
    _format_numeric_columns(sheet, 5, sheet.max_row, {4, 5}, "NT$ #,##0.00")
    _finish_sheet(sheet)


def _write_stock_summary(
    sheet: Worksheet,
    payload: dict[str, Any],
    *,
    generated_at: datetime | None,
) -> None:
    _setup_sheet(sheet, "個股摘要", generated_at=generated_at, last_column=6)
    profile = payload.get("profile") or {}
    summary = payload.get("summary") or {}
    quote = payload.get("quote") or {}
    brief = payload.get("brief") or {}
    valuation = payload.get("valuation") or {}
    suitability = valuation.get("suitability") or {}
    relative = valuation.get("relative") or {}

    rows = [
        ["代號", profile.get("stock_id")],
        ["名稱", profile.get("short_name") or profile.get("name")],
        ["公司全名", profile.get("name")],
        ["市場", profile.get("market")],
        ["產業", profile.get("industry_label")],
        ["最近收盤", summary.get("latest_close")],
        ["收盤資料日", summary.get("end_date")],
        ["盤中參考", quote.get("display_price")],
        ["估值適用性", suitability.get("state_label")],
        ["主要觀察方法", (suitability.get("recommended") or {}).get("primary_label")],
        ["相對估值狀態", relative.get("headline")],
        ["白話公司說明", brief.get("company_sentence")],
        ["估值白話說明", brief.get("valuation_sentence")],
        ["風險標籤", "、".join(brief.get("risk_tags") or [])],
    ]
    _write_table(sheet, 4, ["項目", "內容"], rows)
    _finish_sheet(sheet)


def _write_structure_fingerprint(
    sheet: Worksheet,
    payload: dict[str, Any],
    *,
    generated_at: datetime | None,
) -> None:
    _setup_sheet(sheet, "結構指紋", generated_at=generated_at, last_column=8)
    structure = payload.get("structure") or {}
    sufficiency = structure.get("sufficiency") if isinstance(structure.get("sufficiency"), dict) else {}
    top_rows = [
        ["標題", structure.get("title") or "結構指紋"],
        ["說明", structure.get("subtitle") or "這檔股票現在的性格（結構描述，非預測）"],
        ["資料日", structure.get("as_of_date")],
        ["視窗", structure.get("window")],
        ["樣本筆數", sufficiency.get("bars_available")],
        ["資料充足度", _structure_grade_label(sufficiency.get("grade"))],
        ["提醒", structure.get("disclaimer") or "結構描述工具 · 描述現在 · 不預測未來 · 非投資建議"],
    ]
    _write_table(sheet, 4, ["項目", "內容"], top_rows)

    rows: list[list[Any]] = []
    for item in structure.get("dimensions") or []:
        if not isinstance(item, dict):
            continue
        rows.append(
            [
                item.get("label") or item.get("key"),
                "鎖定" if item.get("locked") else ("可用" if item.get("available") else "資料不足"),
                _structure_grade_label(item.get("grade")),
                _structure_bar_text(item),
                item.get("summary"),
                item.get("forbidden"),
                item.get("overlap_note"),
                _structure_raw_text(item),
            ]
        )
    _write_table(
        sheet,
        13,
        ["維度", "狀態", "等級", "格數", "摘要", "不要這樣解讀", "差異說明", "原始讀數"],
        rows or [["尚無結構指紋資料", None, None, None, None, None, None, None]],
    )
    _finish_sheet(sheet)


def _write_prices(
    sheet: Worksheet,
    payload: dict[str, Any],
    *,
    generated_at: datetime | None,
) -> None:
    _setup_sheet(sheet, "日線資料", generated_at=generated_at, last_column=8)
    headers = ["日期", "開盤", "最高", "最低", "收盤", "成交量", "漲跌", "來源"]
    rows = [
        [
            item.get("date"),
            item.get("open"),
            item.get("high"),
            item.get("low"),
            item.get("close"),
            item.get("volume"),
            item.get("change"),
            item.get("source"),
        ]
        for item in payload.get("prices") or []
    ]
    _write_table(sheet, 4, headers, rows or [["尚無日線資料", None, None, None, None, None, None, None]])
    _format_numeric_columns(sheet, 5, sheet.max_row, {2, 3, 4, 5, 7}, "#,##0.00")
    _format_numeric_columns(sheet, 5, sheet.max_row, {6}, "#,##0")
    _finish_sheet(sheet)


def _write_chart_annotations(
    sheet: Worksheet,
    payload: dict[str, Any],
    *,
    generated_at: datetime | None,
) -> None:
    _setup_sheet(sheet, "圖表標註", generated_at=generated_at, last_column=8)
    headers = ["類型", "起點日期", "起點價格", "終點日期", "終點價格", "文字", "顏色", "更新時間"]
    rows = [
        [
            _annotation_kind_label(item.get("kind")),
            item.get("anchor_date"),
            item.get("anchor_price"),
            item.get("anchor_date2"),
            item.get("anchor_price2"),
            item.get("text"),
            item.get("color"),
            item.get("updated_at"),
        ]
        for item in payload.get("annotations") or []
        if isinstance(item, dict)
    ]
    _write_table(sheet, 4, headers, rows or [["尚無圖表標註", None, None, None, None, None, None, None]])
    _format_numeric_columns(sheet, 5, sheet.max_row, {3, 5}, "#,##0.00")
    _finish_sheet(sheet)


def _write_dividends(
    sheet: Worksheet,
    payload: dict[str, Any],
    *,
    generated_at: datetime | None,
) -> None:
    _setup_sheet(sheet, "股利資料", generated_at=generated_at, last_column=9)
    headers = ["年度", "期間", "現金股利", "股票股利", "狀態", "董事會日", "資料日", "來源", "註記"]
    rows = [
        [
            item.get("year"),
            item.get("period"),
            item.get("cash_dividend"),
            item.get("stock_dividend"),
            item.get("status"),
            item.get("board_date"),
            item.get("source_updated_at"),
            item.get("source"),
            item.get("note"),
        ]
        for item in payload.get("dividends") or []
    ]
    _write_table(sheet, 4, headers, rows or [["尚無股利資料", None, None, None, None, None, None, None, None]])
    _format_numeric_columns(sheet, 5, sheet.max_row, {3, 4}, "#,##0.####")
    _finish_sheet(sheet)


def _write_valuation(
    sheet: Worksheet,
    payload: dict[str, Any],
    *,
    generated_at: datetime | None,
) -> None:
    _setup_sheet(sheet, "估值情境", generated_at=generated_at, last_column=7)
    valuation = payload.get("valuation") or {}
    suitability = valuation.get("suitability") or {}
    relative = valuation.get("relative") or {}
    historical = valuation.get("historical_yield") or {}
    dividend_summary = valuation.get("dividend_summary") or {}

    top_rows = [
        ["估值適用性", suitability.get("state_label")],
        ["資料信心", suitability.get("data_confidence_label")],
        ["公司類型", suitability.get("company_type_label")],
        ["主要方法", (suitability.get("recommended") or {}).get("primary_label")],
        ["股利資料年數", dividend_summary.get("years")],
        ["平均現金股利", dividend_summary.get("average_cash_dividend")],
        ["股票股利口徑", dividend_summary.get("stock_dividend_scope_note")],
        ["相對估值說明", relative.get("headline")],
    ]
    _write_table(sheet, 4, ["項目", "內容"], top_rows)

    estimates = []
    for item in historical.get("estimates") or []:
        estimates.append(
            [
                _dividend_scenario_label(item.get("scenario")),
                _as_percent_ratio(item.get("target_yield_percent")),
                item.get("price"),
            ]
        )
    _write_table(
        sheet,
        14,
        ["股利情境", "目標殖利率", "反推價格"],
        estimates or [["尚無歷史殖利率情境", None, None]],
    )
    _format_numeric_columns(sheet, 15, sheet.max_row, {2}, "0.00%")
    _format_numeric_columns(sheet, 15, sheet.max_row, {3}, "#,##0.00")

    row = sheet.max_row + 3
    methods = relative.get("methods") or []
    multiple_rows: list[list[Any]] = []
    for method in methods:
        for estimate in method.get("estimates") or []:
            multiple_rows.append(
                [
                    method.get("title"),
                    estimate.get("label"),
                    estimate.get("multiple"),
                    estimate.get("price"),
                    method.get("warning"),
                ]
            )
    _write_table(
        sheet,
        row,
        ["敏感度方法", "情境", "倍數", "反推價格", "提醒"],
        multiple_rows or [["尚無 PE/PB 敏感度資料", None, None, None, None]],
    )
    _finish_sheet(sheet)


def _write_health_report(
    sheet: Worksheet,
    payload: dict[str, Any],
    *,
    generated_at: datetime | None,
) -> None:
    _setup_sheet(sheet, "白話健檢", generated_at=generated_at, last_column=4)
    report = payload.get("report") or {}
    rows = []
    for section in report.get("sections") or []:
        rows.append(
            [
                section.get("title"),
                section.get("tone"),
                section.get("summary"),
                " / ".join(section.get("details") or []),
            ]
        )
    _write_table(sheet, 4, ["面向", "語氣", "摘要", "細節"], rows or [["尚無白話健檢", None, None, None]])
    _finish_sheet(sheet)


def _write_monthly_revenue(
    sheet: Worksheet,
    payload: dict[str, Any],
    *,
    generated_at: datetime | None,
) -> None:
    _setup_sheet(sheet, "月營收", generated_at=generated_at, last_column=7)
    headers = ["月份", "當月營收", "月增率", "年增率", "累計營收", "累計年增率", "資料日"]
    rows = [
        [
            item.get("year_month"),
            item.get("current_month_revenue"),
            _as_percent_ratio(item.get("mom_percent")),
            _as_percent_ratio(item.get("yoy_percent")),
            item.get("cumulative_revenue"),
            _as_percent_ratio(item.get("cumulative_yoy_percent")),
            item.get("source_updated_at"),
        ]
        for item in payload.get("monthly_revenues") or []
    ]
    _write_table(sheet, 4, headers, rows or [["尚無月營收資料", None, None, None, None, None, None]])
    _format_numeric_columns(sheet, 5, sheet.max_row, {2, 5}, "#,##0")
    _format_numeric_columns(sheet, 5, sheet.max_row, {3, 4, 6}, "0.00%")
    _finish_sheet(sheet)


def _write_financials(
    sheet: Worksheet,
    payload: dict[str, Any],
    *,
    generated_at: datetime | None,
) -> None:
    _setup_sheet(sheet, "財報摘要", generated_at=generated_at, last_column=10)
    headers = ["季度", "營收", "毛利", "營益", "淨利", "EPS", "毛利率", "營益率", "淨利率", "ROE"]
    rows = [
        [
            item.get("quarter_label"),
            item.get("revenue"),
            item.get("gross_profit"),
            item.get("operating_income"),
            item.get("net_income"),
            item.get("eps"),
            _as_percent_ratio(item.get("gross_margin_percent")),
            _as_percent_ratio(item.get("operating_margin_percent")),
            _as_percent_ratio(item.get("net_margin_percent")),
            _as_percent_ratio(item.get("roe_percent")),
        ]
        for item in payload.get("financial_statements") or []
    ]
    _write_table(sheet, 4, headers, rows or [["尚無財報資料", None, None, None, None, None, None, None, None, None]])
    _format_numeric_columns(sheet, 5, sheet.max_row, {2, 3, 4, 5}, "#,##0")
    _format_numeric_columns(sheet, 5, sheet.max_row, {6}, "0.00")
    _format_numeric_columns(sheet, 5, sheet.max_row, {7, 8, 9, 10}, "0.00%")
    _finish_sheet(sheet)


def _structure_bar_text(item: dict[str, Any]) -> str:
    if item.get("locked"):
        return "需市場資料"
    if not item.get("available"):
        return "資料不足"
    return f"{item.get('bar_level')}/{item.get('bar_max') or 5}"


def _structure_grade_label(value: Any) -> str:
    if value == "high":
        return "充足"
    if value == "medium":
        return "僅供參考"
    if value == "low":
        return "偏少"
    if value == "locked":
        return "鎖定"
    return "不足"


def _structure_raw_text(item: dict[str, Any]) -> str:
    if item.get("grade") in ("low", "insufficient"):
        return ""
    raw = item.get("raw") if isinstance(item.get("raw"), dict) else {}
    parts: list[str] = []
    for key, value in raw.items():
        number = _to_float(value)
        parts.append(f"{key}={number:.4f}" if number is not None else f"{key}={value}")
    return "、".join(parts)


def _write_notes(
    sheet: Worksheet,
    rows: list[list[Any]],
    *,
    generated_at: datetime | None,
) -> None:
    _setup_sheet(sheet, "說明", generated_at=generated_at, last_column=2)
    _write_table(sheet, 4, ["項目", "說明"], rows)
    _finish_sheet(sheet)


def _setup_sheet(
    sheet: Worksheet,
    title: str,
    *,
    generated_at: datetime | None,
    last_column: int,
) -> None:
    generated_at = generated_at or datetime.now()
    last = get_column_letter(max(last_column, 2))
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells(f"A1:{last}1")
    sheet["A1"] = (
        f"資料日期：{generated_at.strftime('%Y-%m-%d %H:%M')}｜"
        f"{SOURCE_NOTE}｜{DISCLAIMER}"
    )
    sheet["A1"].fill = PatternFill("solid", fgColor=NOTE_FILL)
    sheet["A1"].font = Font(color="536271", size=10)
    sheet["A1"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[1].height = 32

    sheet.merge_cells(f"A2:{last}2")
    sheet["A2"] = title
    sheet["A2"].fill = PatternFill("solid", fgColor=BRAND_FILL)
    sheet["A2"].font = Font(color="FFFFFF", bold=True, size=14)
    sheet["A2"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[2].height = 26


def _write_table(
    sheet: Worksheet,
    start_row: int,
    headers: list[str],
    rows: list[list[Any]],
) -> None:
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(start_row, col, header)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(bold=True, color="203040")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(bottom=Side(style="thin", color=BORDER_COLOR))
    for cell in sheet[start_row]:
        cell.border = border

    for row_offset, row in enumerate(rows, start=1):
        for col, value in enumerate(row, start=1):
            cell = sheet.cell(start_row + row_offset, col, _excel_cell_value(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _finish_sheet(sheet: Worksheet) -> None:
    sheet.freeze_panes = "A4"
    thin = Side(style="thin", color=BORDER_COLOR)
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            cell.border = Border(bottom=thin)
    widths: dict[int, int] = {}
    for row in sheet.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), min(len(value) + 3, 44))
    for col, width in widths.items():
        sheet.column_dimensions[get_column_letter(col)].width = max(11, width)


def _format_numeric_columns(
    sheet: Worksheet,
    start_row: int,
    end_row: int,
    columns: set[int],
    number_format: str,
) -> None:
    for row in range(start_row, end_row + 1):
        for col in columns:
            sheet.cell(row, col).number_format = number_format


def _save_workbook(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _as_percent_ratio(value: Any) -> float | None:
    number = _to_float(value)
    return number / 100 if number is not None else None


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _excel_cell_value(value: Any) -> Any:
    if isinstance(value, list):
        return "、".join(str(item) for item in value)
    if isinstance(value, dict):
        return json_dumps(value)
    return value


def json_dumps(value: Any) -> str:
    import json

    return json.dumps(value, ensure_ascii=False, sort_keys=True)


def _annotation_kind_label(value: Any) -> str:
    labels = {
        "note": "文字",
        "line": "線段",
        "arrow": "箭頭",
        "range": "區間",
        "gap": "缺口",
    }
    return labels.get(str(value or "note"), str(value or "文字"))


def _dividend_scenario_label(scenario: Any) -> str:
    if scenario == "high_yield":
        return "高殖利率情境"
    if scenario == "average_yield":
        return "平均殖利率情境"
    if scenario == "low_yield":
        return "低殖利率情境"
    return str(scenario or "情境")
