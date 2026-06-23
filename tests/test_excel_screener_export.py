import io
import unittest

from openpyxl import load_workbook

from app.exporters.excel import build_screener_workbook_bytes

SAMPLE_PAYLOAD = {
    "generated_at": "2026-06-16T01:00:00+00:00",
    "targets": {"cheap": 6.25, "fair": 5.0, "expensive": 3.125},
    "summary": {
        "rows": 2,
        "available_rows": 1,
        "below_cheap_high_conf_rows": 1,
        "yield_normal_rows": 1,
        "yield_trap_rows": 1,
    },
    "items": [
        {
            "stock_id": "2412",
            "short_name": "中華電",
            "name": "中華電信",
            "market": "TWSE",
            "price_date": "2026-06-13",
            "latest_close": 125.5,
            "previous_close": 124.8,
            "day_change": 0.7,
            "day_change_percent": 0.56,
            "volume": 12000,
            "average_cash_dividend": 4.7,
            "current_yield_percent": 3.75,
            "data_years": 5,
            "cheap_price": 75.2,
            "fair_price": 94.0,
            "expensive_price": 150.4,
            "difference_percent": 66.9,
            "suitability_state": "applicable",
            "company_type_label": "成熟高息股",
            "suitability_reasons": [],
            "yield_trap": False,
            "yield_trap_reason": "",
            "status": "5 年資料",
            "confidence_label": "信心較高",
        },
        {
            "stock_id": "3576",
            "short_name": "聯合再生",
            "name": "聯合再生能源",
            "market": "TWSE",
            "price_date": "2026-06-13",
            "latest_close": 21.0,
            "previous_close": 19.8,
            "day_change": 1.2,
            "day_change_percent": 6.06,
            "volume": 50000,
            "average_cash_dividend": 0.1,
            "current_yield_percent": 0.48,
            "data_years": 2,
            "cheap_price": 1.6,
            "fair_price": 2.0,
            "expensive_price": 3.2,
            "difference_percent": 1212.5,
            "suitability_state": "not_applicable",
            "company_type_label": "景氣循環股·太陽能",
            "suitability_reasons": ["yield_too_low", "cyclical", "loss_history"],
            "yield_trap": True,
            "yield_trap_reason": "one_off_dividend",
            "status": "2 年資料",
            "confidence_label": "低信心",
        },
    ],
}


class ScreenerExcelExportTest(unittest.TestCase):
    def test_workbook_opens_and_has_sheets(self):
        content = build_screener_workbook_bytes(SAMPLE_PAYLOAD)
        self.assertIsInstance(content, bytes)
        self.assertGreater(len(content), 0)
        workbook = load_workbook(io.BytesIO(content))
        self.assertIn("雷達總表", workbook.sheetnames)
        self.assertIn("傳統估價表", workbook.sheetnames)
        self.assertIn("說明", workbook.sheetnames)

    def test_main_sheet_has_no_misleading_words(self):
        # 主表（雷達總表）與說明維持中性命名；同事版「傳統估價表」是刻意的例外，不在此檢查。
        content = build_screener_workbook_bytes(SAMPLE_PAYLOAD)
        workbook = load_workbook(io.BytesIO(content))
        text_blob = []
        for name in ("雷達總表", "說明"):
            for row in workbook[name].iter_rows(values_only=True):
                for value in row:
                    if isinstance(value, str):
                        text_blob.append(value)
        joined = "\n".join(text_blob)
        for forbidden in ("便宜價", "合理價", "昂貴價", "低估", "高估", "該買"):
            self.assertNotIn(forbidden, joined)

    def test_colleague_sheet_has_legacy_columns(self):
        content = build_screener_workbook_bytes(SAMPLE_PAYLOAD)
        workbook = load_workbook(io.BytesIO(content))
        joined = "\n".join(
            str(value)
            for row in workbook["傳統估價表"].iter_rows(values_only=True)
            for value in row
            if value is not None
        )
        for want in ("便宜價(買價)", "合理價", "昂貴價(賣價)", "5年平均", "差異", "情境參考價"):
            self.assertIn(want, joined)

    def test_reason_and_trap_labels_translated(self):
        content = build_screener_workbook_bytes(SAMPLE_PAYLOAD)
        workbook = load_workbook(io.BytesIO(content))
        sheet = workbook["雷達總表"]
        joined = "\n".join(
            str(value)
            for row in sheet.iter_rows(values_only=True)
            for value in row
            if value is not None
        )
        self.assertIn("不適合", joined)
        self.assertIn("景氣循環股", joined)
        self.assertIn("含一次性高股利", joined)


if __name__ == "__main__":
    unittest.main()
