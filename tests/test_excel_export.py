from __future__ import annotations

import json
from io import BytesIO
import unittest

from openpyxl import load_workbook

from app.exporters.excel import (
    build_portfolio_workbook_bytes,
    build_stock_workbook_bytes,
)


class ExcelExportTests(unittest.TestCase):
    def test_portfolio_export_contains_summary_and_transactions(self) -> None:
        payload = {
            "summary": {
                "positions_count": 1,
                "transactions_count": 2,
                "total_cost_basis": 100000,
                "total_market_value": 108000,
                "total_unrealized_pnl": 8000,
                "total_unrealized_return_percent": 8,
                "realized_pnl": 1200,
                "cost_method": "移動平均成本法",
                "price_basis": "最近收盤價",
            },
            "positions": [
                {
                    "stock_id": "2330",
                    "profile": {"short_name": "台積電"},
                    "shares": 100,
                    "average_cost": 1000,
                    "cost_basis": 100000,
                    "latest_close": 1080,
                    "latest_close_date": "2026-06-16",
                    "market_value": 108000,
                    "unrealized_pnl": 8000,
                    "unrealized_return_percent": 8,
                }
            ],
            "transactions": [
                {
                    "trade_date": "2026-06-01",
                    "stock_id": "2330",
                    "profile": {"short_name": "台積電"},
                    "side": "buy",
                    "shares": 100,
                    "price": 1000,
                    "fee": 20,
                    "tax": 0,
                    "note": "測試",
                }
            ],
            "performance": {
                "total_cash_dividends": 2000,
                "total_return_amount": 10000,
                "total_return_percent": 10,
                "xirr_percent": 11.5,
                "cash_dividend_events": [
                    {
                        "ex_date": "2026-06-15",
                        "stock_id": "2330",
                        "shares": 100,
                        "cash_dividend_per_share": 20,
                        "cash_amount": 2000,
                        "source": "TWSE_TWT49U",
                    }
                ],
                "benchmark": {
                    "total_return_percent": 6,
                },
            },
        }

        workbook = load_workbook(BytesIO(build_portfolio_workbook_bytes(payload)))

        self.assertEqual(workbook.sheetnames, ["持倉總覽", "交易紀錄", "配息紀錄", "說明"])
        self.assertIn("非投資建議", workbook["持倉總覽"]["A1"].value)
        self.assertEqual(workbook["持倉總覽"]["A16"].value, "2330")
        self.assertEqual(workbook["交易紀錄"]["B5"].value, "2330")
        self.assertEqual(workbook["配息紀錄"]["B5"].value, "2330")
        self.assertIn("尚未納入", workbook["說明"]["A6"].value)

    def test_stock_export_uses_neutral_valuation_language(self) -> None:
        payload = {
            "profile": {
                "stock_id": "2330",
                "short_name": "台積電",
                "name": "台灣積體電路製造股份有限公司",
                "market": "TWSE",
                "industry_label": "半導體業",
            },
            "summary": {"latest_close": 1080, "end_date": "2026-06-16"},
            "quote": {"display_price": 1085},
            "brief": {
                "company_sentence": "台積電屬於半導體業。",
                "valuation_sentence": "股利不是主要報酬來源，股利法不適合當主尺。",
                "risk_tags": ["成長取向"],
            },
            "valuation": {
                "suitability": {
                    "state_label": "股利法參考性低",
                    "data_confidence_label": "信心中等",
                    "company_type_label": "成長型",
                    "recommended": {"primary_label": "本益比（PE）敏感度"},
                },
                "relative": {
                    "headline": "這是 what-if，不是預測。",
                    "methods": [
                        {
                            "title": "本益比（PE）倍數敏感度",
                            "warning": "PE 敏感度不判斷價格高低。",
                            "estimates": [
                                {"label": "市場少付 20%", "multiple": 24, "price": 864},
                                {"label": "目前倍數", "multiple": 30, "price": 1080},
                            ],
                        }
                    ],
                },
                "dividend_summary": {
                    "years": 5,
                    "average_cash_dividend": 15,
                },
                "historical_yield": {
                    "estimates": [
                        {
                            "scenario": "high_yield",
                            "target_yield_percent": 6.25,
                            "price": 240,
                        }
                    ]
                },
            },
            "structure": {
                "available": True,
                "as_of_date": "2026-06-16",
                "window": 250,
                "title": "結構指紋",
                "subtitle": "這檔股票現在的性格（結構描述，非預測）",
                "disclaimer": "結構描述工具 · 描述現在 · 不預測未來 · 非投資建議",
                "sufficiency": {"bars_available": 320, "grade": "high"},
                "dimensions": [
                    {
                        "key": "memory",
                        "label": "延續性",
                        "available": True,
                        "bar_level": 3,
                        "bar_max": 5,
                        "grade": "high",
                        "summary": "H=0.58：延續性偏高，傾向延續近期行為。",
                        "forbidden": "不得解讀為後續方向承諾或必然反轉；延續性描述自相關結構，不含方向。",
                        "overlap_note": "與圖上趨勢強度不同。",
                        "raw": {"hurst_dfa": 0.58},
                    }
                ],
            },
            "prices": [{"date": "2026-06-16", "open": 1000, "high": 1100, "low": 990, "close": 1080, "volume": 1, "change": 80, "source": "test"}],
            "annotations": [
                {
                    "kind": "gap",
                    "anchor_date": "2026-06-16",
                    "anchor_price": 1080,
                    "text": "觀察缺口",
                    "color": "#2C5475",
                    "updated_at": "2026-06-16T08:30:00",
                }
            ],
            "dividends": [{"year": 115, "period": "年度", "cash_dividend": 15, "stock_dividend": 0, "status": "除息", "board_date": "2026-06-01", "source_updated_at": "2026-06-01", "source": "test"}],
            "report": {"sections": [{"title": "價格位階", "tone": "neutral", "summary": "中性", "details": ["不等於可買"]}]},
            "monthly_revenues": [],
            "financial_statements": [],
        }

        content = build_stock_workbook_bytes(payload)
        workbook = load_workbook(BytesIO(content))
        text = json.dumps(
            [
                cell.value
                for sheet in workbook.worksheets
                for row in sheet.iter_rows()
                for cell in row
                if cell.value is not None
            ],
            ensure_ascii=False,
        )

        self.assertIn("估值情境", workbook.sheetnames)
        self.assertIn("圖表標註", workbook.sheetnames)
        self.assertIn("結構指紋", workbook.sheetnames)
        self.assertIn("非投資建議", workbook["個股摘要"]["A1"].value)
        self.assertEqual(workbook["結構指紋"]["A14"].value, "延續性")
        self.assertEqual(workbook["結構指紋"]["D14"].value, "3/5")
        self.assertEqual(workbook["估值情境"]["A15"].value, "高殖利率情境")
        self.assertAlmostEqual(workbook["估值情境"]["B15"].value, 0.0625)
        self.assertEqual(workbook["圖表標註"]["A5"].value, "缺口")
        self.assertEqual(workbook["圖表標註"]["F5"].value, "觀察缺口")
        self.assertNotIn("便宜價", text)
        self.assertNotIn("合理價", text)
        self.assertNotIn("昂貴價", text)


if __name__ == "__main__":
    unittest.main()
